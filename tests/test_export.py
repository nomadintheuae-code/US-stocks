"""Phase 8 CSV/Excel Export tests (no network)."""
import csv
import pytest
from pathlib import Path
from unittest.mock import patch

from engines.export import (
    export_csv,
    export_excel,
    run_export,
    _flatten_candidate,
    DEFAULT_COLUMNS,
    EXTENDED_COLUMNS,
    ALL_COLUMNS,
)


# ---------------------------------------------------------------------------
# Test data
# ---------------------------------------------------------------------------

SAMPLE_CANDIDATES = [
    {
        "ticker": "AAPL", "status": "ACTION", "price": 195.50,
        "entry": 196.00, "stop": 190.00, "target": 210.00, "shares": 50,
        "rs": 85, "pf": 2.5, "sector": "Technology",
        "vcp": {"score": 72, "signals": ["tightness", "volume_dry"]},
        "analyst_target": 210.0, "analyst_upside": 7.5,
        "recommendation": "Buy", "insider_alert": False,
        "earnings_warning": None, "bb_squeeze": False,
        "candle_bias": "bullish", "fib_nearest": 192.0,
    },
    {
        "ticker": "MSFT", "status": "ACTION", "price": 420.00,
        "entry": 421.00, "stop": 410.00, "target": 450.00, "shares": 20,
        "rs": 78, "pf": 1.8, "sector": "Technology",
        "vcp": {"score": 65, "signals": ["ma_alignment"]},
        "analyst_target": 450.0, "analyst_upside": 7.1,
        "recommendation": "Strong Buy", "insider_alert": True,
        "earnings_warning": "Earnings in 5d", "bb_squeeze": True,
        "candle_bias": "neutral", "fib_nearest": None,
    },
    {
        "ticker": "JPM", "status": "WAIT", "price": 195.00,
        "entry": 196.00, "stop": 188.00, "target": 220.00, "shares": 50,
        "rs": 72, "pf": 1.3, "sector": "Financials",
        "vcp": {"score": 58, "signals": []},
        "analyst_target": None, "analyst_upside": None,
        "recommendation": None, "insider_alert": False,
        "earnings_warning": None, "bb_squeeze": False,
        "candle_bias": None, "fib_nearest": 193.0,
    },
]


# ---------------------------------------------------------------------------
# Flatten tests
# ---------------------------------------------------------------------------

class TestFlattenCandidate:

    def test_basic_columns(self):
        row = _flatten_candidate(SAMPLE_CANDIDATES[0], DEFAULT_COLUMNS)
        assert row["ticker"] == "AAPL"
        assert row["status"] == "ACTION"
        assert row["price"] == 195.50

    def test_vcp_score(self):
        row = _flatten_candidate(SAMPLE_CANDIDATES[0], ["vcp_score"])
        assert row["vcp_score"] == 72

    def test_vcp_signals(self):
        row = _flatten_candidate(SAMPLE_CANDIDATES[0], ["vcp_signals"])
        assert row["vcp_signals"] == "tightness, volume_dry"

    def test_empty_vcp_signals(self):
        row = _flatten_candidate(SAMPLE_CANDIDATES[2], ["vcp_signals"])
        assert row["vcp_signals"] == ""

    def test_missing_field(self):
        row = _flatten_candidate(SAMPLE_CANDIDATES[0], ["nonexistent"])
        assert row["nonexistent"] is None


# ---------------------------------------------------------------------------
# CSV export tests
# ---------------------------------------------------------------------------

class TestExportCSV:

    def test_csv_basic(self, tmp_path):
        out = tmp_path / "test.csv"
        result = export_csv(SAMPLE_CANDIDATES[:1], out)
        assert result.exists()
        with open(result, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["ticker"] == "AAPL"

    def test_csv_multiple_rows(self, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(SAMPLE_CANDIDATES, out)
        with open(out, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 3

    def test_csv_custom_columns(self, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(SAMPLE_CANDIDATES[:1], out, columns=["ticker", "price"])
        with open(out, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert list(rows[0].keys()) == ["ticker", "price"]

    def test_csv_empty(self, tmp_path):
        out = tmp_path / "test.csv"
        result = export_csv([], out)
        assert result.exists()
        with open(result, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert rows == []

    def test_csv_creates_parent_dir(self, tmp_path):
        out = tmp_path / "sub" / "dir" / "test.csv"
        export_csv(SAMPLE_CANDIDATES[:1], out)
        assert out.exists()

    def test_csv_full_columns(self, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(SAMPLE_CANDIDATES[:1], out, columns=ALL_COLUMNS)
        with open(out, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows[0]) == len(ALL_COLUMNS)


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------

class TestExportExcel:

    def test_excel_basic(self, tmp_path):
        out = tmp_path / "test.xlsx"
        result = export_excel(SAMPLE_CANDIDATES[:1], out)
        assert result.exists()
        assert out.suffix == ".xlsx"

    def test_excel_with_watchlist(self, tmp_path):
        out = tmp_path / "test.xlsx"
        watchlist = [SAMPLE_CANDIDATES[2]]
        result = export_excel(
            SAMPLE_CANDIDATES[:2], out,
            include_watchlist=True, watchlist=watchlist,
        )
        assert result.exists()
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Watchlist" in wb.sheetnames
        ws = wb["Watchlist"]
        assert ws.cell(row=1, column=1).value == "ticker"
        assert ws.cell(row=2, column=1).value == "JPM"

    def test_excel_no_watchlist(self, tmp_path):
        out = tmp_path / "test.xlsx"
        export_excel(SAMPLE_CANDIDATES[:1], out, include_watchlist=False)
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Scan Results"

    def test_excel_creates_parent_dir(self, tmp_path):
        out = tmp_path / "sub" / "dir" / "test.xlsx"
        export_excel(SAMPLE_CANDIDATES[:1], out)
        assert out.exists()

    def test_excel_format_options(self, tmp_path):
        out = tmp_path / "test.xlsx"
        export_excel(
            SAMPLE_CANDIDATES[:1], out,
            format_options={"header_font": True, "auto_width": True, "freeze_panes": True},
        )
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.freeze_panes == "A2"
        assert ws.cell(row=1, column=1).font.bold is True

    def test_excel_custom_columns(self, tmp_path):
        out = tmp_path / "test.xlsx"
        export_excel(SAMPLE_CANDIDATES[:1], out, columns=["ticker", "rs"])
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "ticker"
        assert ws.cell(row=1, column=2).value == "rs"
        assert ws.cell(row=2, column=1).value == "AAPL"


# ---------------------------------------------------------------------------
# run_export tests
# ---------------------------------------------------------------------------

class TestRunExport:

    def test_run_export_csv_and_excel(self, tmp_path):
        selected = SAMPLE_CANDIDATES[:2]
        full = SAMPLE_CANDIDATES
        watchlist = [SAMPLE_CANDIDATES[2]]
        result = run_export(
            selected=selected,
            qualified_full=full,
            watchlist_wait=watchlist,
            output_dir=tmp_path,
            csv_enabled=True,
            excel_enabled=True,
        )
        assert "csv" in result
        assert "csv_full" in result
        assert "excel" in result
        assert result["csv"].exists()
        assert result["csv_full"].exists()
        assert result["excel"].exists()

    def test_run_export_csv_only(self, tmp_path):
        result = run_export(
            selected=SAMPLE_CANDIDATES[:1],
            qualified_full=SAMPLE_CANDIDATES[:1],
            output_dir=tmp_path,
            csv_enabled=True,
            excel_enabled=False,
        )
        assert "csv" in result
        assert "excel" not in result

    def test_run_export_excel_only(self, tmp_path):
        result = run_export(
            selected=SAMPLE_CANDIDATES[:1],
            qualified_full=SAMPLE_CANDIDATES[:1],
            output_dir=tmp_path,
            csv_enabled=False,
            excel_enabled=True,
        )
        assert "csv" not in result
        assert "excel" in result

    def test_run_export_empty(self, tmp_path):
        result = run_export(
            selected=[],
            qualified_full=[],
            output_dir=tmp_path,
        )
        # Should still create files (empty)
        assert result["csv"].exists()
        assert result["excel"].exists()


# ---------------------------------------------------------------------------
# Config tests
# ---------------------------------------------------------------------------

class TestExportConfig:

    def test_default_disabled(self):
        from sentinel.config import ExportConfig
        cfg = ExportConfig()
        assert cfg.enabled is False
        assert cfg.csv is True
        assert cfg.excel is True
        assert cfg.include_watchlist is True
        assert cfg.columns is None

    def test_in_main_config(self):
        from sentinel.config import Config
        cfg = Config()
        assert hasattr(cfg, "export")
        assert cfg.export.enabled is False


__all__ = [
    "TestFlattenCandidate",
    "TestExportCSV",
    "TestExportExcel",
    "TestRunExport",
    "TestExportConfig",
]

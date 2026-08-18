"""CSV/Excel Export Engine for SENTINEL PRO.

Provides scan result export to CSV and Excel formats.  Opt-in via
``export.enabled`` in config.yaml.  When disabled, no export files
are created and the scanner's default behavior is unchanged.

Features:
- CSV export with configurable columns and encoding
- Excel export with optional formatting (header styling, auto-width)
- Separate exports for ACTION, watchlist, and full qualified lists
- Filename includes date for easy tracking
"""
import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


# ───────────────────────────────────────────────────────────────────────
# Default column definitions
# ───────────────────────────────────────────────────────────────────────

DEFAULT_COLUMNS = [
    "ticker", "status", "price", "entry", "stop", "target", "shares",
    "rs", "pf", "sector",
]

EXTENDED_COLUMNS = DEFAULT_COLUMNS + [
    "vcp_score", "vcp_signals", "analyst_target", "analyst_upside",
    "recommendation", "insider_alert", "earnings_warning",
    "bb_squeeze", "candle_bias", "fib_nearest",
]

ALL_COLUMNS = EXTENDED_COLUMNS + [
    "analyst_count", "short_ratio", "short_pct",
    "insider_pct", "institution_pct", "pe_forward", "revenue_growth",
    "price_vs_pivot_pct",
]


def _flatten_candidate(c: Dict[str, Any], columns: List[str]) -> Dict[str, Any]:
    """Flatten a candidate dict into a flat row for CSV/Excel.

    Handles nested VCP dict → flat columns (vcp_score, vcp_signals).
    """
    row: Dict[str, Any] = {}
    for col in columns:
        if col == "vcp_score":
            vcp = c.get("vcp", {})
            row[col] = vcp.get("score") if isinstance(vcp, dict) else None
        elif col == "vcp_signals":
            vcp = c.get("vcp", {})
            signals = vcp.get("signals", []) if isinstance(vcp, dict) else []
            row[col] = ", ".join(signals) if signals else ""
        elif col == "price_vs_pivot_pct":
            price = c.get("price")
            pivot = c.get("price")  # fallback
            entry = c.get("entry")
            if price and entry and entry > 0:
                row[col] = round((price - entry) / entry * 100, 2)
            else:
                row[col] = None
        else:
            row[col] = c.get(col)
    return row


# ───────────────────────────────────────────────────────────────────────
# CSV Export
# ───────────────────────────────────────────────────────────────────────

def export_csv(
    candidates: List[Dict],
    output_path: Path,
    columns: Optional[List[str]] = None,
    encoding: str = "utf-8-sig",
) -> Path:
    """Export candidates to a CSV file.

    Args:
        candidates: list of candidate dicts (qualified, selected, etc.)
        output_path: file path to write to
        columns: column list (defaults to DEFAULT_COLUMNS)
        encoding: file encoding (default utf-8-sig for Excel compat)

    Returns:
        Path to the written file.
    """
    columns = columns or DEFAULT_COLUMNS
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = [_flatten_candidate(c, columns) for c in candidates]

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    logger.info("CSV export: %d rows → %s", len(rows), output_path)
    return output_path


# ───────────────────────────────────────────────────────────────────────
# Excel Export
# ───────────────────────────────────────────────────────────────────────

def export_excel(
    candidates: List[Dict],
    output_path: Path,
    columns: Optional[List[str]] = None,
    sheet_name: str = "Scan Results",
    include_watchlist: bool = False,
    watchlist: Optional[List[Dict]] = None,
    format_options: Optional[Dict[str, Any]] = None,
) -> Path:
    """Export candidates to an Excel file (.xlsx).

    Args:
        candidates: list of candidate dicts (ACTION/selected)
        output_path: file path to write to
        columns: column list (defaults to ALL_COLUMNS)
        sheet_name: name of the main sheet
        include_watchlist: if True, add a second sheet with watchlist items
        watchlist: watchlist candidates (required if include_watchlist=True)
        format_options: dict with optional keys:
            - header_font: bold header (default True)
            - auto_width: auto-adjust column widths (default True)
            - freeze_panes: freeze header row (default True)

    Returns:
        Path to the written file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    columns = columns or ALL_COLUMNS
    fmt = format_options or {}
    do_header_font = fmt.get("header_font", True)
    do_auto_width = fmt.get("auto_width", True)
    do_freeze = fmt.get("freeze_panes", True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── Main sheet ──
    ws = wb.active
    ws.title = sheet_name
    _write_sheet(ws, candidates, columns, do_header_font, do_auto_width)
    if do_freeze:
        ws.freeze_panes = "A2"

    # ── Watchlist sheet ──
    if include_watchlist and watchlist:
        ws2 = wb.create_sheet(title="Watchlist")
        _write_sheet(ws2, watchlist, columns, do_header_font, do_auto_width)
        if do_freeze:
            ws2.freeze_panes = "A2"

    wb.save(str(output_path))
    logger.info("Excel export: %d rows → %s", len(candidates), output_path)
    return output_path


def _write_sheet(
    ws,
    candidates: List[Dict],
    columns: List[str],
    do_header_font: bool,
    do_auto_width: bool,
) -> None:
    """Write data to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        if do_header_font:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, candidate in enumerate(candidates, 2):
        flat = _flatten_candidate(candidate, columns)
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=flat.get(col_name))

    # Auto-width
    if do_auto_width:
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row_idx in range(2, len(candidates) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else get_column_letter(col_idx)].width = min(max_len + 2, 30)


# ───────────────────────────────────────────────────────────────────────
# High-level export runner
# ───────────────────────────────────────────────────────────────────────

def run_export(
    selected: List[Dict],
    qualified_full: List[Dict],
    watchlist_wait: Optional[List[Dict]] = None,
    output_dir: Optional[Path] = None,
    csv_enabled: bool = True,
    excel_enabled: bool = True,
    columns: Optional[List[str]] = None,
) -> Dict[str, Path]:
    """Run full export (CSV + Excel) for scan results.

    Args:
        selected: ACTION candidates
        qualified_full: all qualified candidates
        watchlist_wait: WAIT candidates for watchlist sheet
        output_dir: output directory (defaults to ./results)
        csv_enabled: whether to export CSV
        excel_enabled: whether to export Excel
        columns: column list for CSV (defaults to DEFAULT_COLUMNS)

    Returns:
        Dict of {format: Path} for each exported file.
    """
    output_dir = output_dir or Path("./results")
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    exported: Dict[str, Path] = {}

    if csv_enabled:
        csv_path = export_csv(
            selected,
            output_dir / f"{date_str}_action.csv",
            columns=columns,
        )
        exported["csv"] = csv_path

        # Also export full qualified list
        csv_full = export_csv(
            qualified_full,
            output_dir / f"{date_str}_full.csv",
            columns=columns,
        )
        exported["csv_full"] = csv_full

    if excel_enabled:
        excel_path = export_excel(
            selected,
            output_dir / f"{date_str}_scan.xlsx",
            include_watchlist=bool(watchlist_wait),
            watchlist=watchlist_wait,
        )
        exported["excel"] = excel_path

    return exported


__all__ = [
    "export_csv",
    "export_excel",
    "run_export",
    "DEFAULT_COLUMNS",
    "EXTENDED_COLUMNS",
    "ALL_COLUMNS",
]
